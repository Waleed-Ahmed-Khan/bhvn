import pandas as pd 
import common.helper as helper
from openpyxl import load_workbook
import streamlit as st
from st_aggrid import AgGrid
import numpy as np
from datetime import timedelta


def upload_sales(df, username):
    sheets = ['Data 2021','Data 2022'] 
    wb = load_workbook(df, read_only=True)
    check = all ( sheet in wb.sheetnames for sheet in  sheets )
    print(check)
    if check:
        logistics_df = pd.concat([pd.read_excel(df, sheet_name=s, header=3).assign(sheet_name=s) for s in sheets]).\
        drop(columns=['Year', 'Quarter', 'Container Number']).rename(columns={'sheet_name':"Year"})
        logistics_df = logistics_df.dropna(subset=['Customer PO Number','Blue Stock PO number', 'BHVN PO Number']).reset_index(drop=True).\
            rename(columns={'Order Creation Date/ Order Received Date (PO Month)':'Month','Date confirmed by BH/Expected Handover date(EHD)':'EHD',
            'Customer handover date/Customer Required Date(CHD)':'CHD','Actual handover date/Actual shipdate (AHD)/Ex-BH Date':'AHD', 'Ship Date/Vessel sailing date':'Ship Date'})
        logistics_df['Year'] = logistics_df['Year'].apply(lambda x : x[5:])
        logistics_df = logistics_df.astype({"Month":str,"Sales Month":str, "EHD":str, "Inv. Date":str, "Ship Date":str, "CHD":str,
                                    "AHD":str, "ETA":str,"Date of Booking Placed":str, "Date of Booking Confirmation":str})
        st.success(f"Great work {str.title(username)}! Data has been validated Successfully 🎉")

        try:
            logistics_df.to_csv("sales_data.csv", index= False)
            # cosmos = MongoDBManagement(conn_string = helper.COSMOS_CONN_STR)
            # cosmos.saveDataFrameIntoCollection("bhvnSales", "bhvndb", logistics_df)
            # st.success(f"Well done {str.title(username)}! Data has been Uploaded Successfully 🎊")
            # #st.warning(f'Dear {str.title(username)}, data is already available in the database so instead of UPLOAD, please UPDATE the data')
            # userManagement.app_logger(username, "Add Sales Data")
        except Exception as e:
            raise e
            #st.error(f"Oh {str.title(username)}, data could not be uploaded in the database Please contact admin in case the problem persists. We are sorry 😌")
    else:
        st.error(f"Sorry {str.title(username)} Requreid Sheets {sheets} is not available in the workbook")

def update_sales(df, db_name, collection_name, username):
    sheets = ['Data 2022'] 
    wb = load_workbook(df, read_only=True)
    check = all ( sheet in wb.sheetnames for sheet in  sheets )
    print(check)
    if check:
        logistics_df = pd.concat([pd.read_excel(df, sheet_name=s, header=3).assign(sheet_name=s) for s in sheets]).\
        drop(columns=['Year', 'Quarter', 'Container Number']).rename(columns={'sheet_name':"Year"})
        logistics_df = logistics_df.dropna(subset=['Customer PO Number','Blue Stock PO number', 'BHVN PO Number']).reset_index(drop=True).\
            rename(columns={'Order Creation Date/ Order Received Date (PO Month)':'Month','Date confirmed by BH/Expected Handover date(EHD)':'EHD',
            'Customer handover date/Customer Required Date(CHD)':'CHD','Actual handover date/Actual shipdate (AHD)/Ex-BH Date':'AHD', 'Ship Date/Vessel sailing date':'Ship Date'})
        logistics_df['Year'] = logistics_df['Year'].apply(lambda x : x[5:])
        logistics_df = logistics_df.astype({"Month":str,"Sales Month":str, "EHD":str, "Inv. Date":str, "Ship Date":str, "CHD":str,
                                    "AHD":str, "ETA":str,"Date of Booking Placed":str, "Date of Booking Confirmation":str})
        st.success(f"Great work {str.title(username)}! Data has been validated Successfully 🎉")
        
        try:
            logistics_df.to_csv("sales_data.csv", index=False)
            # cosmos = MongoDBManagement(conn_string = helper.COSMOS_CONN_STR)
            # query = {"Year":"2022"}
            # cosmos.deleteRecord(db_name, collection_name, query)
            # userManagement.app_logger(username, "Updated Sales Data")
            # cosmos.saveDataFrameIntoCollection("bhvnSales", "bhvndb", logistics_df)
            # st.success(f"Well done {str.title(username)}! Data has been Updated Successfully 🎊")
        except Exception as e:
            raise e
            #st.error(f"Sorry {str.title(username)}, we could not update the data, please contact admin if the problem persists")
    else: 
        st.error(f"Sorry {str.title(username)} Requreid Sheets {sheets} is not available in the workbook")

@st.cache(suppress_st_warning=True, show_spinner =False, ttl = 60)
def load_sales_data(query): 
    # cosmos = MongoDBManagement(conn_string = helper.COSMOS_CONN_STR)
    # df = cosmos.getDataFrameOfCollection("bhvndb", "bhvnSales", query)
    df = pd.read_csv('sales_data.csv')
    df = df[df['Customer']=="DCL"]
    return df

def get_sales_data_l():
    query = {"Customer":"DCL"}
    df = load_sales_data(query= query) # to apply customer filtering, pass the query variable here
    df = df.astype({"Month":"datetime64","Sales Month":"datetime64", "EHD":"datetime64", "Inv. Date":"datetime64", "Ship Date":"datetime64", "CHD":"datetime64",
                                    "AHD":"datetime64", "ETA":"datetime64","Date of Booking Placed":"datetime64", "Date of Booking Confirmation":"datetime64"})
    return df

def data_processing_sales_l(df, start_date, end_date, po_selection, customer_selection, brand_selection):
    mask =  (df['Sales Month'].astype('datetime64').dt.date.between(start_date, end_date)) & (df['Customer PO Number'].isin(po_selection)) & (df['Customer'].isin(customer_selection)) & (df['Brand'].isin(brand_selection))
    df = df.loc[mask]
    df['EHD_inc'] = df["EHD"] + timedelta(days=7)   
    df['CHD_inc'] = df["CHD"] + timedelta(days=4)
    df['Delay EHD']  = (df['AHD'] - df['EHD_inc']).apply(lambda x: x.days)
    df['Delay CHD'] = (df['AHD'] - df['CHD_inc']).apply(lambda x: x.days)
    df['month_number'] = df['Sales Month'].dt.month
    df.sort_values("month_number", ascending=True, inplace = True)
    df['SalesMonth'] = df['Sales Month'].apply(lambda x : x.strftime("%b"))

    df['Booking Confirmation Days']  = (df['Date of Booking Confirmation'] - df['Date of Booking Placed']).apply(lambda x: x.days)
    df['Total Value (US$)'] = df['Qty'] * df['Unit price (US$)']
    df['Sailing Delay'] = (df['Ship Date'] - df['Date of Booking Confirmation']).apply(lambda x: x.days)
    df['Sailing Days'] = (df['ETA'] - df['Ship Date']).apply(lambda x: x.days)
    monthly_data = df.groupby(by='SalesMonth').agg({"Qty": np.sum, "Delay CHD": np.mean, "Delay EHD":np.mean, "Booking Confirmation Days":np.mean, 
                    "Sailing Delay":np.mean, "Sailing Days": np.mean, "month_number":pd.Series.mode}).sort_values(by='month_number', ascending=True).reset_index()
    
    customer_wise = df.groupby(by='Customer').agg({"Qty": np.sum, "Delay CHD": np.mean, "Delay EHD":np.mean, "Booking Confirmation Days":np.mean, 
                    "Sailing Delay":np.mean, "Sailing Days": np.mean}).reset_index()
    monthly_data = monthly_data.rename(columns={"Qty": "TotalQty"})
   
    
    #df = df.astype({'Delay EHD':int, 'Delay CHD':int})
    df['Delay EHD'] = df['Delay EHD'].values.astype(int)
    df['Delay CHD'] = df['Delay CHD'].values.astype(int)
    df['Delay/Ontime_EHD'] = np.where(df['Delay EHD'].isnull(),np.nan,
                             np.where(df['Delay EHD'] < 0,"Advance",
                             np.where((df['Delay EHD'] <= 7)&(df['Delay EHD'].values >= 0), "OnTime", "Delayed")))
    df['Delay/Ontime_CHD'] = np.where(df['Delay CHD'].isnull(),np.nan,
                             np.where(df['Delay CHD'] == 0,"OnTime",
                             np.where((df['Delay CHD'] > 4), "OnTime", "Delayed")))
 

    '''
    np.where(df['Delay CHD'].isnull(),np.nan,
                                np.where(df['Delay CHD'].values == 0,"OnTime",
                                np.where((df['Delay CHD'].values <= 7)&(df["CHD"] < "2022-04-30"), "OnTime",
                                np.where((df['Delay CHD'].values <= 4)&(df["CHD"] >= "2022-04-30"), "OnTime", "Delayed"))))

    '''
    
    def get_hot(df, monthly_data, delay_col, hot_title = "HOT", cols_to_rename = None):
        ontime_qty = df[(df[delay_col]=='OnTime') | (df[delay_col]=="Advance")].groupby(by='SalesMonth')\
                    .agg({"Qty": np.sum, "month_number":pd.Series.mode}).reset_index()
        delay_qty = df[df[delay_col]=='Delayed'].groupby(by='SalesMonth').agg({"Qty": np.sum}).reset_index()
        
        monthly_data = pd.merge(monthly_data, ontime_qty[['SalesMonth', 'Qty']], on = 'SalesMonth', how = "left").rename(columns={'Qty':cols_to_rename[0]})
        monthly_data = pd.merge(monthly_data, delay_qty[['SalesMonth', 'Qty']], on = 'SalesMonth', how = "left").rename(columns={'Qty':cols_to_rename[1]})
        
        monthly_data[hot_title] = np.where((monthly_data[cols_to_rename[1]].isnull()) & (monthly_data[cols_to_rename[0]].isnull()),np.nan,
                             np.where((monthly_data[cols_to_rename[1]].notnull()) & (monthly_data[cols_to_rename[0]].notnull()) ,  (monthly_data[cols_to_rename[0]]/monthly_data['TotalQty'])*100,
                             np.where((monthly_data[cols_to_rename[1]].notnull()) & (monthly_data[cols_to_rename[0]].isnull()), 0,
                             np.where((monthly_data[cols_to_rename[1]].isnull()) & (monthly_data[cols_to_rename[0]].notnull()), 100 , np.nan))))
        
        return monthly_data

    
    monthly_data = get_hot(df = df[df['Year']==2021], monthly_data=monthly_data, delay_col = 'Delay/Ontime_CHD', hot_title= 'HOT_CHD_2021', cols_to_rename = ['QtyOntimeCHD2021', 'QtyDelayedCHD2021'] )
    
    monthly_data = get_hot(df = df[df['Year']==2021], monthly_data=monthly_data, delay_col = 'Delay/Ontime_EHD', hot_title= 'HOT_EHD_2021', cols_to_rename = ['QtyOntimeEHD2021', 'QtyDelayedEHD2021'] )
    monthly_data = get_hot(df = df[df['Year']==2022], monthly_data=monthly_data, delay_col = 'Delay/Ontime_CHD', hot_title= 'HOT_CHD_2022', cols_to_rename = ['QtyOntimeCHD2022', 'QtyDelayedCHD2022'] )
    monthly_data = get_hot(df = df[df['Year']==2022], monthly_data=monthly_data, delay_col = 'Delay/Ontime_EHD', hot_title= 'HOT_EHD_2022', cols_to_rename = ['QtyOntimeEHD2022', 'QtyDelayedEHD2022'] )

    ontime_qty_ehd = df[(df['Delay/Ontime_EHD']=='OnTime') | (df['Delay/Ontime_EHD']=="Advance")].groupby(by='SalesMonth')\
                .agg({"Qty": np.sum, "month_number":pd.Series.mode}).reset_index()
    delay_qty_ehd = df[df['Delay/Ontime_EHD']=='Delayed'].groupby(by='SalesMonth').agg({"Qty": np.sum}).reset_index()
    #hot_chd = pd.merge(ontime_qty_chd, delay_qty_chd, on="Month", how='left', suffixes=('OntimeCHD', 'DelayedCHD')).drop('month_number', axis = 1)
    monthly_data = pd.merge(monthly_data, ontime_qty_ehd[['SalesMonth', 'Qty']], on = 'SalesMonth', how = "left").rename(columns={'Qty':'QtyOntimeEHD'}).drop('month_number', axis = 1)
    monthly_data = pd.merge(monthly_data, delay_qty_ehd[['SalesMonth', 'Qty']], on = 'SalesMonth', how = "left").rename(columns={'Qty':'QtyDelayedEHD'})

    monthly_data['HOT_EHD'] = np.where(pd.notnull(monthly_data['QtyDelayedEHD']), monthly_data['QtyOntimeEHD']/monthly_data['TotalQty'], 1)


    ontime_qty_chd = df[(df['Delay/Ontime_CHD']=='OnTime') | (df['Delay/Ontime_CHD']=="Advance")].groupby(by='SalesMonth')\
                .agg({"Qty": np.sum, "month_number":pd.Series.mode}).reset_index()
    delay_qty_chd = df[df['Delay/Ontime_CHD']=='Delayed'].groupby(by='SalesMonth').agg({"Qty": np.sum}).reset_index()
    
    monthly_data = pd.merge(monthly_data, ontime_qty_chd[['SalesMonth', 'Qty']], on = 'SalesMonth', how = "left").rename(columns={'Qty':'QtyOntimeCHD'})
    monthly_data = pd.merge(monthly_data, delay_qty_chd[['SalesMonth', 'Qty']], on = 'SalesMonth', how = "left").rename(columns={'Qty':'QtyDelayedCHD'})
    monthly_data['HOT_CHD'] = np.where(pd.notnull(monthly_data['QtyDelayedCHD']), monthly_data['QtyOntimeCHD']/monthly_data['TotalQty'], 1)
    return df, monthly_data, customer_wise

@st.cache(suppress_st_warning=True, show_spinner =False, ttl = 5000)
def load_sales_invoice_data():
    query = '''
    SELECT det.customer_ponumber,inv.bhvnpo, inv.custumInvoiceNo, det.invoiceid, inv.invoicedate, inv.shipdate,
        inv.booking_pdate,inv.booking_rdate, inv.shipcost,inv.AdvancePayment, inv.freight,
        inv.cosignee, inv.cosig_address, inv.brand, inv.bankname, inv.container, 
        inv.vesseldate, inv.vesselno, det.itemid, det.itemdescription, det.priceach, det.qty, 
        det.amount, det.discountPercent, det.exchange_rate, det.price_InCurrency,det.lotnumber, det.Cartons, 
        det.modelcode, det.hscode, det.imancode, det.container
    FROM tblinvoice inv
    LEFT JOIN tblinvoicedetail det ON
        det.invoiceid = inv.invoicenumber;
    '''
    inv = helper.load_data(query=query)
    inv.rename(columns = {'customer_ponumber':'custponumber'}, inplace =True)
    inv = inv[inv['custponumber'] !='']
    query = '''
    SELECT tblarorder.ponumber, tblarorder.id AS SalesOrder, tblarorder.ProductionNo, tblarorderdetail.description, 
		   tblarorderdetail.itemid, tblitems.size, itemqtyorder, tblarorder.shipTo_comp_name AS custname, lotnumber, tblarorder.confirmdate,
           tblarorder.duedate,tblarorder.ex_bhvn_date, tblarorder.bhvnorderentrydate, tblarorder.orderdate
    FROM tblarorderdetail 
    LEFT JOIN  tblarorder ON 
           tblarorder.id = tblarorderdetail.orderid 
    LEFT JOIN tblitems ON
           tblitems.itemid = tblarorderdetail.itemid
    '''
    order = helper.load_data(query=query)
    complete_sales = pd.merge(inv, order, how = 'left',
                             left_on=['custponumber', 'itemid'],
                             right_on = ['ponumber', 'itemid']
                             )
    #complete_sales = complete_sales.astype({'invoicedate': "datetime64[ns]" , 'shipdate': "datetime64[ns]"})
    #complete_sales['pomonth_number'] = pd.to_datetime(complete_sales['orderdate'])
    time_stamp = ['invoicedate', 'shipdate', 'orderdate', 'confirmdate', 'duedate', 'ex_bhvn_date', 'bhvnorderentrydate', 'booking_rdate','booking_pdate']
    for col in time_stamp:
        complete_sales[col] = [pd.to_datetime(x, infer_datetime_format=True) if pd.notnull(x) else x for x in complete_sales[col]]
    
    non_time_stamp = [col for col in complete_sales.columns if col not in time_stamp]
    for col in non_time_stamp:
        complete_sales[col] = [np.nan if pd.isnull(x) or len(str(x))<1 else x for x in complete_sales[col]]
    complete_sales.rename(columns = {'qty': 'shipped_qty', 'itemqtyorder':'order_qty', 'price_InCurrency':'Price(USD)'}, inplace=True)
    
    complete_sales['pomonth_number'] = complete_sales['orderdate'].dt.month
    complete_sales['salesmonth_number'] = complete_sales['shipdate'].dt.month
    complete_sales['year'] = complete_sales['shipdate'].dt.year.astype(int)
    complete_sales['pomonth'] = complete_sales['orderdate'].dt.month_name().str.slice(stop=3)
    complete_sales['salesmonth'] = complete_sales['shipdate'].dt.month_name().str.slice(stop=3)
    complete_sales['target_lead_time'] = (complete_sales['duedate'] - complete_sales['orderdate']).dt.days
    complete_sales['actual_lead_time'] = np.where(((complete_sales['shipdate'].isnull()) | (complete_sales['orderdate'].isnull())), 0, (complete_sales['shipdate'] - complete_sales['orderdate']).dt.days) 
    complete_sales['delay_days'] = complete_sales['actual_lead_time'] - complete_sales['target_lead_time']
    complete_sales['delay_days'] = np.where(complete_sales['delay_days'] < 0, 0, complete_sales['delay_days'])
    complete_sales['booking_conf_time'] = np.where((complete_sales['booking_rdate'].isnull()) | (complete_sales['booking_pdate'].isnull()), 0,  (complete_sales['booking_rdate'] - complete_sales['booking_pdate']).dt.days)
    complete_sales['Delayed Qty'] = np.where(complete_sales['delay_days'] <= 0, 0 , complete_sales['shipped_qty'])
    complete_sales['percent_qty_delayed'] = (complete_sales['Delayed Qty']/complete_sales['shipped_qty'])*100
    complete_sales['HOT'] =  ((complete_sales['shipped_qty'] - complete_sales['Delayed Qty']) / complete_sales['shipped_qty']) *100
    
    return complete_sales

#@st.cache(suppress_st_warning=True, allow_output_mutation=True, ttl = 3)
def preprocess_sales(complete_df, start_date, end_date, year, po_selection, customer_selection,inv_selection):
    mask =  (complete_df['invoicedate'].astype('datetime64').dt.date.between(start_date, end_date)) & (complete_df['year'] == year) & (complete_df['ponumber'].isin(po_selection)) & (complete_df['custname'].isin(customer_selection)) & (complete_df['custumInvoiceNo'].isin(inv_selection))
    complete_df = complete_df[mask]
    return complete_df